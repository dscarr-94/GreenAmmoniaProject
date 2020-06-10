import yaml
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, colors
import time
import sys
import math
from tqdm import tqdm
YELLOW_HIGHLIGHT = PatternFill(start_color='ffff00',
				   end_color='ffff00',
				   fill_type='solid')
GREEN_HIGHLIGHT = PatternFill(start_color='90ee90',
				   end_color='90ee90',
				   fill_type='solid')
ORANGE_HIGHLIGHT = PatternFill(start_color='ffa500',
				   end_color='ffa500',
				   fill_type='solid')

#Round to two sig figs 
def round_val(num):
	if(num < .001 and num > .001):
		return 0.0
	else:
		return num

#Retrieves data from the YAML file and returns a dictionary with the vals.
def get_config_variables():
	with open('config.yaml') as f:
		data = yaml.load(f, Loader=yaml.FullLoader)
		return data

#Copies the current worksheet into a new worksheet, returns worksheet
def copy_worksheet(workbook, sheetName):
	source = workbook.active
	target = workbook.copy_worksheet(source)
	target.title = sheetName
	return target

#NOTE: This only works for the first column!!
#Returns the row (in column 1) equal to the key string, returns int
def find_row_with_key(worksheet, key):
	rowNum = 0 #loop through and find row for key
	for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1, values_only=True):
		rowNum += 1
		for cell in row:
			if(cell != None and isinstance(cell, str)):
				if(key in cell):
				   return rowNum
	return 0

#Removes all rows below "Mass Flows"
def remove_rows_below(worksheet):
	rowNum = 0 #loop through and find row for mass flows
	massFlowArr = []
	for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1, values_only=True):
		rowNum += 1
		for cell in row:
			if(cell == "Mass Flows"):
			   massFlowArr.append(rowNum) #More than one intance of mass flow in sheet
	worksheet.delete_rows(massFlowArr[0] + 1, worksheet.max_row) #Delete from first mass flow cell down

#Remove all the rows that have zero values in them throughout
def remove_zero_rows(worksheet):
	maxRow = worksheet.max_row
	maxCol = worksheet.max_column
	rowNum = 3
	for row in worksheet.iter_rows(min_col=3, max_col=maxCol, min_row=3, max_row=maxRow, values_only=True):
		allString = False
		zeroFlag = False #Assume all zero
		for item in row:
			if(isinstance(item, (float, int))): #Is a number?
				if(item != None):
					if(item != 0):
						zeroFlag = True
		for item in row:
			if (isinstance(item, str)):
				allString = True
		if(zeroFlag == False and allString == False):
			worksheet.delete_rows(rowNum, 1)
		rowNum += 1

#Adds title to the worksheet
def addTitle(worksheet, title):
	worksheet.insert_rows(1)
	worksheet['A1'] = title

#Adds the Section In/Out values and unmerges the cells
def addInOutRows(worksheet):
	worksheet.insert_rows(2)
	worksheet['A2'] = "Section In"
	worksheet.insert_rows(3)
	worksheet['A3'] = "Section Out"
	worksheet.merge_cells('A2:CO2') #Bug fix, must merge and unmerge to just unmerge
	worksheet.unmerge_cells('A2:CO2') 

#Performs and writes the entropy calculations to the streams worksheet
def entropyCalculations(worksheet):
	rowIdx = 0
	maxRow = worksheet.max_row
	maxCol = worksheet.max_column
	flag = False
	for row in worksheet.iter_rows(min_row=1, min_col=1,
		max_col=1, max_row=maxRow, values_only=True):
		rowIdx += 1
		for item in row:
			if(item == "Enthalpy Flow"):
				flag = True #found the one row I needed
				newRow = rowIdx + 1
				worksheet.insert_rows(newRow, 2) #Create new row below Enthalpy
				titleCell = "A" + str(newRow)
				unitCell = "B" + str(newRow)
				worksheet[titleCell] = "Entropy Flow" #Add new name for row
				worksheet[unitCell] = "kW/K"
				titleCell = "A" + str(newRow + 1)
				unitCell = "B" + str(newRow + 1)
				worksheet[titleCell] = "Exergy Flow" #Add new name for row
				worksheet[unitCell] = "MW"
		if(flag):
			break
	conversionFactor = 3600 * 1000
	conversionFactorSeconds = 1000
	stream_molar_flow = find_row_with_key(worksheet, "Mole Flows")
	stream_molar_flow_units = "B" + str(stream_molar_flow)
	stream_molar_entropy = find_row_with_key(worksheet, "Molar Entropy")
	stream_molar_entropy_units = "B" + str(stream_molar_entropy)

	for row_cells in worksheet.iter_rows(min_row=newRow, max_row=newRow,
		min_col=3, max_col=maxCol):
		for cell in row_cells:
			firstValue = worksheet[str(cell.column_letter) + str(stream_molar_entropy)]
			secondValue = worksheet[str(cell.column_letter) + str(stream_molar_flow)]
			if(firstValue.value != None and secondValue.value != None):
				if(worksheet[stream_molar_flow_units].value == "kmol/hr" and 
					worksheet[stream_molar_entropy_units].value == "J/kmol-K"):
					calculatedValue = (firstValue.value * secondValue.value)/conversionFactor
				elif(worksheet[stream_molar_flow_units].value == "kmol/sec" and 
					worksheet[stream_molar_entropy_units].value == "J/kmol-K"):
					calculatedValue = (firstValue.value * secondValue.value)/conversionFactorSeconds
				else:
					calculatedValue = 1
					print("Unit error in calculating Entropy Flow, required: \n")
					print("Mole Flow Rate: kmol/hr")
					print("Entropy Mixture: J/kmol-K")
				cell.value = calculatedValue
			#Exergy Calculations
			firstValue = worksheet[str(cell.column_letter) + str(find_row_with_key(worksheet, "Enthalpy Flow"))]
			secondValue = worksheet[str(cell.column_letter) + str(find_row_with_key(worksheet, "Entropy Flow"))]
			if(firstValue.value != None and secondValue.value != None):
				calculatedValue = firstValue.value - (0.3 * secondValue.value)
				cell.offset(row=1).value = calculatedValue

	worksheet.delete_rows(find_row_with_key(worksheet, "Description"), 1)

#Returns column letter of first blank after to and from rows
def find_blank(worksheet):
	to_idx = find_row_with_key(worksheet, "To")
	from_idx = find_row_with_key(worksheet, "From")

	for col in worksheet.iter_cols(min_row=to_idx, max_row=to_idx, min_col=3, max_col=worksheet.max_column):
		for cell in col:
			to_cell = str(cell.column_letter) + str(to_idx)
			from_cell = str(cell.column_letter) + str(from_idx)
			if worksheet[to_cell].value == None and worksheet[from_cell].value == None:
				return cell.column_letter
				#worksheet.delete_cols(cell.column_letter,1)

#Removes columns that have both in and out values in respective columns
def removeColumns(worksheet):
	to_idx = find_row_with_key(worksheet, "To")
	from_idx = find_row_with_key(worksheet, "From")
	x = find_blank(worksheet)
	if(x == None):
		endCol = worksheet.max_column+1
	else:
		endCol = column_index_from_string(x)

	delArray = []
	for col in worksheet.iter_cols(min_row=to_idx, max_row=to_idx, min_col=3, max_col=endCol):
		for cell in col:
			to_cell = str(cell.column_letter) + str(to_idx)
			from_cell = str(cell.column_letter) + str(from_idx)
			if worksheet[to_cell].value != None and worksheet[from_cell].value != None:
				delArray.append(column_index_from_string(cell.column_letter))

	for i in reversed(delArray): #Must delete in reversed array as it changes excel sheet during manipulation
		worksheet.delete_cols(i,1)
	
#Writes In and Our to the correct columns
def addInOutValues(worksheet):
	to_idx = find_row_with_key(worksheet, "To")
	from_idx = find_row_with_key(worksheet, "From")
	x = find_blank(worksheet)
	if(x == None):
		endCol = worksheet.max_column
	else:
		endCol = column_index_from_string(x)
	for col in worksheet.iter_cols(min_row=to_idx, max_row=to_idx, min_col=3, max_col=endCol):
		for cell in col:
			write_cell = str(cell.column_letter) + str(1)
			to_cell = str(cell.column_letter) + str(to_idx)
			from_cell = str(cell.column_letter) + str(from_idx)
			if worksheet[to_cell].value != None:
				worksheet[write_cell] = "In"
			if worksheet[from_cell].value != None:
				worksheet[write_cell] = "Out"

	worksheet.delete_cols(endCol, worksheet.max_column) 

#Calculates and writes the sum values as well as mass flows
def calculate_balance(worksheet):
	enthalpy_flow = find_row_with_key(worksheet, "Enthalpy Flow")
	enthalpy_sum = 0 #Initialize counter
	for col in worksheet.iter_cols(min_row=enthalpy_flow, max_row=enthalpy_flow, min_col=3, max_col=worksheet.max_column):
		lastCol = col
		for cell in col:
			in_out_row = str(cell.column_letter) + "1"
			if worksheet[in_out_row].value == "In": #Check if the first row in this col has In or Out
				enthalpy_sum += cell.value 			#If In, Add to the sum
			if worksheet[in_out_row].value == "Out":
				enthalpy_sum -= cell.value 			#If out, Subtract from sum
	lastCol[0].offset(column=1).value = enthalpy_sum
	lastCol[0].offset(column=1).fill = YELLOW_HIGHLIGHT

	entropy_flow = find_row_with_key(worksheet, "Entropy Flow")
	entropy_sum = 0 #Initialize counter	
	for col in worksheet.iter_cols(min_row=entropy_flow, max_row=entropy_flow, min_col=3, max_col=worksheet.max_column):
		lastCol = col
		for cell in col:
			in_out_row = str(cell.column_letter) + "1"
			if worksheet[in_out_row].value == "In": #Check if the first row in this col has In or Out
				entropy_sum += cell.value 			#If In, Add to the sum
				# print("Plus: " + str(cell.value))
			if worksheet[in_out_row].value == "Out":
				entropy_sum -= cell.value 			#If out, Subtract from sum
				# print("Minus: " + str(cell.value))
	lastCol[0].value = entropy_sum * -1 #Bug: Negative vals calculated unsure why
	lastCol[0].fill = YELLOW_HIGHLIGHT

	exergy_flow = find_row_with_key(worksheet, "Exergy Flow")
	exergy_sum = 0 #Initialize counter	
	for col in worksheet.iter_cols(min_row=exergy_flow, max_row=exergy_flow, min_col=3, max_col=worksheet.max_column):
		lastCol = col
		for cell in col:
			in_out_row = str(cell.column_letter) + "1"
			if worksheet[in_out_row].value == "In": #Check if the first row in this col has In or Out
				exergy_sum += cell.value 			#If In, Add to the sum
			if worksheet[in_out_row].value == "Out":
				exergy_sum -= cell.value 			#If out, Subtract from sum
	lastCol[0].value = exergy_sum * -1
	lastCol[0].fill = YELLOW_HIGHLIGHT
	worksheet[(str((lastCol[0]).column_letter) + "1")].value = "Balances" #Convoluted, just add title to cell

	mass_flows = find_row_with_key(worksheet, "Mass Flows")
	mass_sum = 0 #Initialize counter	
	for col in worksheet.iter_cols(min_row=mass_flows, max_row=mass_flows, min_col=3, max_col=worksheet.max_column):
		lastCol = col
		for cell in col:
			in_out_row = str(cell.column_letter) + "1"
			if worksheet[in_out_row].value == "In": #Check if the first row in this col has In or Out
				mass_sum += cell.value 			#If In, Add to the sum
			if worksheet[in_out_row].value == "Out":
				mass_sum -= cell.value 			#If out, Subtract from sum
	if(abs(mass_sum <= 10)):	#Check if the mass_flow sum is greater than 10, if so MB_Error
		lastCol[0].value = mass_sum
		lastCol[0].fill = YELLOW_HIGHLIGHT
		return 1
	else:	
		print("MB_Error, Mass Flow Sum is: " + str(mass_sum))
		print("See cell: " + str(lastCol[0].coordinate))
		return 0

def step_six(worksheet):
	#Lesson, do not use array for deleting rows because they change dynamically per each deletion
	worksheet.delete_rows(1, 2)
	worksheet.delete_rows(find_row_with_key(worksheet, "Maximum Relative Error"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Cost Flow"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "MIXED Substream"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Vapor Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Liquid Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Solid Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Enthalpy"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Entropy"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Density"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Molar Liquid Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Molar Solid Fraction"), 1)
	remove_rows_below(worksheet)
	remove_zero_rows(worksheet)
	enthalpy_flow = find_row_with_key(worksheet, "Enthalpy Flow")
	enthalpy_flow_units = "B" + str(enthalpy_flow)
	if(worksheet[enthalpy_flow_units].value == "Watt"):
		# print("Enthalpy Flow in Watts, converting to MW")
		for col in worksheet.iter_cols(min_col=3, max_col=worksheet.max_column, 
				min_row=enthalpy_flow, max_row=enthalpy_flow):
			if(col[0].value != None):
				newVal = col[0].value / 1000000
				col[0].value = newVal

def add_text(worksheet, title, overall_text):
	worksheet['A1'] = title
	cnt = 0
	name_array= []
	for col in worksheet.iter_rows(min_row=3, max_row=3):
		for cell in col:
			if(cell.value == "Name"):
				name_array.append(cell.col_idx)
	
	for x in name_array:
		cnt = 0
		for col in worksheet.iter_rows(min_row=64, max_row=110, min_col=x, max_col=x):
			for cell in col:
				if (overall_text[cnt] == "Inlet Stream 1 Name"
					or overall_text[cnt] == "Inlet Stream 2 Name"
					or overall_text[cnt] == "Inlet Stream 3 Name"
					or overall_text[cnt] == "Inlet Stream 4 Name"):
					cell.fill = YELLOW_HIGHLIGHT
				if (overall_text[cnt] == "Outlet Stream 1 Name"
					or overall_text[cnt] == "Outlet Stream 2 Name"
					or overall_text[cnt] == "Outlet Stream 3 Name"
					or overall_text[cnt] == "Outlet Stream 4 Name"
					or overall_text[cnt] == "Outlet Stream 5 Name"
					or overall_text[cnt] == "Outlet Stream 6 Name"):
					cell.fill = ORANGE_HIGHLIGHT
				if (overall_text[cnt] == "Mass balance kg/hr"
					or overall_text[cnt] == "Energy Balance MW"
					or overall_text[cnt] == "Entropy Generation kW/K"):
					cell.fill=GREEN_HIGHLIGHT
				cell.value = overall_text[cnt]
			cnt += 1

	for col in worksheet.iter_rows(min_row=3, max_row=3):
		for cell in col:
			if(cell.value != "Name" and cell.value != None):
				temp = cell.value
				thisCell = cell.offset(row=62)
				thisCell.value = temp

	block_name_array = []
	idx = 0
	for col in worksheet.iter_rows(min_row=2, max_row=2):
		for cell in col:
			if(cell.value != None):
				block_name_array.append(cell.value)
	for col in worksheet.iter_cols(min_row=64, max_row=64):
		for cell in col:
			if cell.value == "Block Type":
				cell.offset(column=1).value = block_name_array[idx]
				idx += 1

def copy_worksheet(workbook, sheetName):
	source = workbook.active
	target = workbook.copy_worksheet(source)
	target.title = sheetName
	return target

def step_seven(worksheet,title):
	addTitle(worksheet,title)
	addInOutRows(worksheet)
	entropyCalculations(worksheet)
	freezeCells = worksheet['C7']
	worksheet.freeze_panes = freezeCells

def step_eight(worksheet):
	freezeCells = worksheet['C7']
	worksheet.freeze_panes = freezeCells
	removeColumns(worksheet)
	addInOutValues(worksheet)

def step_nine(worksheet):
	return calculate_balance(worksheet)

#Looks at the to row, returns array of tuples with format as follows:
#	(to-row-name, stream-name)
def prepare_for_overall_inlet(worksheet):
	return_dict = {}
	to_row = find_row_with_key(worksheet,"To")
	stream_name_row = find_row_with_key(worksheet, "Stream Name") - to_row
	for col in worksheet.iter_cols(min_row=to_row, max_row=to_row,min_col=3):
		for cell in col:
			if cell.value != None:
				if(cell.value in return_dict): #Already in dict 
					return_dict[cell.value].append(cell.offset(row=stream_name_row).value);
				else: #New key
					return_dict[cell.value] = [cell.offset(row=stream_name_row).value]
	return return_dict

def prepare_for_overall_inlet_vals(worksheet):
	return_vals_dict = {}
	to_row = find_row_with_key(worksheet,"To")
	mass_flow_row = find_row_with_key(worksheet,"Mass Flows") - to_row
	enthalpy_flow_row = find_row_with_key(worksheet,"Enthalpy Flow") - to_row
	entropy_flow_row = find_row_with_key(worksheet,"Entropy Flow") - to_row
	for col in worksheet.iter_cols(min_row=to_row, max_row=to_row,min_col=3):
		for cell in col:
			if cell.value != None:
				if(cell.value in return_vals_dict):
					return_vals_dict[cell.value].append((cell.offset(row=mass_flow_row).value, 
						cell.offset(row=enthalpy_flow_row).value,
						cell.offset(row=entropy_flow_row).value))
				else: #New key
					return_vals_dict[cell.value] = [(cell.offset(row=mass_flow_row).value, 
						cell.offset(row=enthalpy_flow_row).value,
						cell.offset(row=entropy_flow_row).value)]
	return return_vals_dict

#Looks at the from row, returns array of tuples with format as follows:
#	(from-row-name, stream-name)
def prepare_for_overall_outlet(worksheet):
	return_dict = {}
	from_row = find_row_with_key(worksheet,"From")
	stream_name_row = find_row_with_key(worksheet, "Stream Name") - from_row
	for col in worksheet.iter_cols(min_row=from_row, max_row=from_row,min_col=3):
		for cell in col:
			if cell.value != None:
				if(cell.value in return_dict): #Already in dict 
					return_dict[cell.value].append(cell.offset(row=stream_name_row).value);
				else: #New key
					return_dict[cell.value] = [cell.offset(row=stream_name_row).value]
	return return_dict

def prepare_for_overall_outlet_vals(worksheet):
	return_vals_dict = {}
	from_row = find_row_with_key(worksheet,"From")
	mass_flow_row = find_row_with_key(worksheet,"Mass Flows") - from_row
	enthalpy_flow_row = find_row_with_key(worksheet,"Enthalpy Flow") - from_row
	entropy_flow_row = find_row_with_key(worksheet,"Entropy Flow") - from_row
	for col in worksheet.iter_cols(min_row=from_row, max_row=from_row,min_col=3):
		for cell in col:
			if cell.value != None:
				if(cell.value in return_vals_dict):
					return_vals_dict[cell.value].append((cell.offset(row=mass_flow_row).value, 
						cell.offset(row=enthalpy_flow_row).value,
						cell.offset(row=entropy_flow_row).value))
				else: #New key
					return_vals_dict[cell.value] = [(cell.offset(row=mass_flow_row).value, 
						cell.offset(row=enthalpy_flow_row).value,
						cell.offset(row=entropy_flow_row).value)]
	return return_vals_dict

def step_twelve_inlet(worksheet, inlet_array, inlet_vals_array):
	for col in worksheet.iter_cols(min_row=65, max_row=65,min_col=2):
		for cell in col:
			if cell.value != "Block Name" and cell.value != None:
				for key in inlet_array:
					if cell.value == key:
						arr_len = len(inlet_array[key])
						if(arr_len == 4):
							thisCell = cell.offset(row=13)
							thisCell.value = inlet_array[key][3]
						if(arr_len >= 3):
							thisCell = cell.offset(row=9)
							thisCell.value = inlet_array[key][2]
						if(arr_len >= 2):
							thisCell = cell.offset(row=5)
							thisCell.value = inlet_array[key][1]
						if(arr_len >= 1):
							thisCell = cell.offset(row=1)
							thisCell.value = inlet_array[key][0]
				for key in inlet_vals_array:
					if cell.value == key:
						arr_len = len(inlet_vals_array[key])
						if(arr_len == 4):
							idx = 0
							for val in inlet_vals_array[key][3]:
								thisCell = cell.offset(row=14+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 3):
							idx = 0
							for val in inlet_vals_array[key][2]:
								thisCell = cell.offset(row=10+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 2):
							idx = 0
							for val in inlet_vals_array[key][1]:
								thisCell = cell.offset(row=6+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 1):
							idx = 0
							for val in inlet_vals_array[key][0]:
								thisCell = cell.offset(row=2+idx)
								thisCell.value = val
								idx += 1

def step_twelve_outlet(worksheet, outlet_array, outlet_vals_array):
	for col in worksheet.iter_cols(min_row=65, max_row=65,min_col=2):
		for cell in col:
			if cell.value != "Block Name" and cell.value != None:
				for key in outlet_array:
					if cell.value == key:
						arr_len = len(outlet_array[key])
						if(arr_len == 6):
							thisCell = cell.offset(row=37)
							thisCell.value = outlet_array[key][5]
						if(arr_len >= 5):
							thisCell = cell.offset(row=33)
							thisCell.value = outlet_array[key][4]
						if(arr_len >= 4):
							thisCell = cell.offset(row=29)
							thisCell.value = outlet_array[key][3]
						if(arr_len >= 3):
							thisCell = cell.offset(row=25)
							thisCell.value = outlet_array[key][2]
						if(arr_len >= 2):
							thisCell = cell.offset(row=21)
							thisCell.value = outlet_array[key][1]
						if(arr_len >= 1):
							thisCell = cell.offset(row=17)
							thisCell.value = outlet_array[key][0]
				for key in outlet_vals_array:
					if cell.value == key:
						arr_len = len(outlet_vals_array[key])
						if(arr_len == 6):
							idx = 0
							for val in outlet_vals_array[key][5]:
								thisCell = cell.offset(row=38+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 5):
							idx = 0
							for val in outlet_vals_array[key][4]:
								thisCell = cell.offset(row=34+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 4):
							idx = 0
							for val in outlet_vals_array[key][3]:
								thisCell = cell.offset(row=30+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 3):
							idx = 0
							for val in outlet_vals_array[key][2]:
								thisCell = cell.offset(row=26+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 2):
							idx = 0
							for val in outlet_vals_array[key][1]:
								thisCell = cell.offset(row=22+idx)
								thisCell.value = val
								idx += 1
						if(arr_len >= 1):
							idx = 0
							for val in outlet_vals_array[key][0]:
								thisCell = cell.offset(row=18+idx)
								thisCell.value = val
								idx += 1


def get_block_range(worksheet, keyword):
	start_idx = 0
	end_idx = 0
	for col in worksheet.iter_cols(min_row=2, max_row=2):
		for cell in col:
			if cell.value == keyword:
				start_idx = cell.col_idx
	for col in worksheet.iter_cols(min_row=3, max_row=3, min_col=start_idx):
		for cell in col:
			if cell.value == None:
				end_idx = cell.col_idx - 1
				break
		else:
			continue
		break
	return(start_idx, end_idx)

def heater_move(worksheet):
	Watt_f = 0
	MW_f = 0
	b_range = get_block_range(worksheet, "Heater")
	if(find_row_with_key(worksheet,"Calculated heat duty [MW]") == 0):
		Watt_f = 1
		b_row = find_row_with_key(worksheet,"Calculated heat duty [Watt]")
	else:
		MW_f = 1
		b_row = find_row_with_key(worksheet,"Calculated heat duty [MW]")

	offset = find_row_with_key(worksheet, "Heat MW") - b_row
	for col in worksheet.iter_cols(min_col=b_range[0], max_col=b_range[1], min_row=b_row, max_row=b_row):
		for cell in col:
			if not (isinstance(cell.value, str)) and cell.value != None:
				if(MW_f == 1):
					writeVal = cell.value #Shift it down to right in the Heat MW row 
				elif(Watt_f):
					print("here")
					writeVal = cell.value /1000000#Shift it down to right in the Heat MW row 
				else:
					print("Unknown heat unit, possible error line %d",cell.offset(row=offset))
					writeVal = cell.value #Shift it down to right in the Heat MW row 
				cell.offset(row=offset).value = writeVal

def pump_move(worksheet):
	p_range = get_block_range(worksheet, "Pump")
	p_row_start = 0
	p_row_end = 0
	conv_flag = 0
	for col in worksheet.iter_cols(min_col=p_range[0], max_col=p_range[1]):
		for cell in col:
			if cell.value == "Net work required [MW]":
				p_row_start = cell.row
			if cell.value == "Net work required [Watt]":
				conv_flag = 1
				p_row_start = cell.row
			if cell.value == "Work MW":
				p_row_end = cell.row

	offset = p_row_end - p_row_start
	for col in worksheet.iter_cols(min_col=p_range[0], max_col=p_range[1], min_row=p_row_start, max_row=p_row_start):
		for cell in col:
			if not (isinstance(cell.value, str)) and cell.value != None:
				writeVal = cell.value #Shift it down to right in the Heat MW row 
				if(conv_flag):
					cell.offset(row=offset).value = writeVal / 1000000
				else:
					cell.offset(row=offset).value = writeVal

def compr_move(worksheet):
	c_range = get_block_range(worksheet, "Compr")
	c_row_start = 0
	c_row_end = 0
	conv_flag = 0
	for col in worksheet.iter_cols(min_col=c_range[0], max_col=c_range[1]):
		for cell in col:
			if cell.value == "Net work required [MW]":
				c_row_start = cell.row
			if cell.value == "Net work required [Watt]":
				conv_flag = 1
				c_row_start = cell.row
			if cell.value == "Work MW":
				c_row_end = cell.row

	offset = c_row_end - c_row_start
	for col in worksheet.iter_cols(min_col=c_range[0], max_col=c_range[1], min_row=c_row_start, max_row=c_row_start):
		for cell in col:
			if not (isinstance(cell.value, str)) and cell.value != None:
				writeVal = cell.value #Shift it down to right in the Heat MW row 
				if(conv_flag):
					cell.offset(row=offset).value = writeVal / 1000000
				else:
					cell.offset(row=offset).value = writeVal

def radfrac_move(worksheet):
	r_range = get_block_range(worksheet, "RadFrac")
	val1_array = []
	val2_array = []
	val1_row = 0
	val2_row = 0
	r_row_end = 0
	idx = 0
	cflag1 = 0
	cflag2 = 0
	for col in worksheet.iter_cols(min_col=r_range[0], max_col=r_range[1]):
		for cell in col:
			if cell.value == "Condenser / top stage heat duty [MW]":
				val1_row = cell.row
			if cell.value == "Condenser / top stage heat duty [Watt]":
				cflag1 = 1
				val1_row = cell.row
			if cell.value == "Reboiler heat duty [MW]":
				val2_row = cell.row
			if cell.value == "Reboiler heat duty [Watt]":
				cflag2 = 1
				val2_row = cell.row
			if cell.value == "Heat MW":
				r_row_end = cell.row
	for col in worksheet.iter_cols(min_col=r_range[0], max_col=r_range[1], min_row=val1_row, max_row=val1_row):
		for cell in col:
			if not (isinstance(cell.value, str)) and cell.value != None: #A number
				if(cflag1):
					val1_array.append(cell.value/1000000)
				else:
					val1_array.append(cell.value)
	for col in worksheet.iter_cols(min_col=r_range[0], max_col=r_range[1], min_row=val2_row, max_row=val2_row):
		for cell in col:
			if not (isinstance(cell.value, str)) and cell.value != None: #A number
				if(cflag2):
					val2_array.append(cell.value/1000000)
				else:
					val2_array.append(cell.value)

	for x in range(len(val1_array)): #Can do this only because val1 and val2 arrays same length
		curr = val1_array[x]
		val1_array[x] = val2_array[x] + curr #Mutate in place save space

	for col in worksheet.iter_cols(min_col=r_range[0] + 1, max_col=r_range[1], min_row=r_row_end, max_row=r_row_end):
		for cell in col:
			cell.value = val1_array[idx]
		idx += 1

def step_thirteen(worksheet):
	heater_move(worksheet)
	pump_move(worksheet)
	compr_move(worksheet)
	radfrac_move(worksheet) 

def get_block_arr(worksheet):
	return_arr = []
	start_idx = 0
	end_idx = 0
	flag1 = False
	flag2 = False
	for col in worksheet.iter_cols(min_row=3, max_row=3):
		for cell in col:
			if cell.value == "Name":
				start_idx = cell.col_idx
				flag1 = True
			if cell.value == None:
				end_idx = cell.col_idx
				flag2 = True
			if(flag1 == True and flag2 == True):
				return_arr.append((start_idx+1,end_idx-1))
				start_idx = 0
				end_idx= 0
				flag1 = False
				flag2 = False
	return return_arr

def step_fourteen(worksheet):
	mass_balance_row = find_row_with_key(worksheet, "Mass balance kg/hr")
	block_arr = get_block_arr(worksheet)
	spec_arr = []
	large_arr = []
	sumVal = 0
	for block_range in block_arr:
		for x in range(block_range[0],block_range[1]+1):
			spec_arr.append(x)
		large_arr.append(spec_arr)
		spec_arr = []

	for arr in large_arr:
		current_block_name = (worksheet[(get_column_letter(arr[0]-1) + "2")].value)
		for col in arr:
			sumVal = float(0.0)
			curr = worksheet[(get_column_letter(col) + "67")].value #Inlet1
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "71")].value #Inlet2
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "75")].value #Inlet3
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "79")].value #Inlet4
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "83")].value #Outlet1
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "87")].value #Outlet2
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "91")].value #Outlet3
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "95")].value #Outlet4
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "99")].value #Outlet5
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "103")].value #Outlet6
			if(curr != None): 
				sumVal -= float(curr)
			writeCell = worksheet[(get_column_letter(col) + "108")]
			if(abs(sumVal <= 10)):
				writeCell.value = round_val(sumVal) #Write Cell - i.e Mass Balance 
			else:	
				writeCell.value = round_val(sumVal) #Write Cell - i.e Mass Balance 
				print("MB_Error, Mass Balance is: " + str(sumVal))
				print("See block: " + current_block_name)
				# return 0

	for arr in large_arr:
		current_block_name = (worksheet[(get_column_letter(arr[0]-1) + "2")].value)
		for col in arr:
			sumVal = float(0.0) #Debug
			curr = worksheet[(get_column_letter(col) + "68")].value #Inlet1
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "72")].value #Inlet2
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "76")].value #Inlet3
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "80")].value #Inlet4
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "84")].value #Outlet1
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "88")].value #Outlet2
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "92")].value #Outlet3
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "96")].value #Outlet4
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "100")].value #Outlet5
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "104")].value #Outlet6
			if(curr != None): 
				sumVal -= float(curr)
			curr = worksheet[(get_column_letter(col) + "106")].value #Work MW
			if(curr != None): 
				sumVal += float(curr)
			curr = worksheet[(get_column_letter(col) + "107")].value #Heat MW
			if(curr != None): 
				sumVal += float(curr)
			writeCell = worksheet[(get_column_letter(col) + "109")]
			if(abs(sumVal) <= 1):
				writeCell.value = round_val(sumVal) #Write Cell - i.e Energy Balance 
			else:	
				writeCell.value = round_val(sumVal) #Write Cell - i.e Energy Balance 
				print("MB_Error, Energy Balance is: " + str(sumVal))
				print("See block: " + current_block_name)
				# return 0

	for arr in large_arr:
		current_block_name = (worksheet[(get_column_letter(arr[0]-1) + "2")].value)
		for col in arr:
			sumVal = float(0.0)
			curr = worksheet[(get_column_letter(col) + "85")].value #Outlet1
			if(curr != None): 
				sumVal += curr
			curr = worksheet[(get_column_letter(col) + "89")].value #Outlet2
			if(curr != None): 
				sumVal += curr
			curr = worksheet[(get_column_letter(col) + "93")].value #Outlet3
			if(curr != None): 
				sumVal += curr
			curr = worksheet[(get_column_letter(col) + "97")].value #Outlet4
			if(curr != None): 
				sumVal += curr
			curr = worksheet[(get_column_letter(col) + "101")].value #Outlet5
			if(curr != None): 
				sumVal += curr
			curr = worksheet[(get_column_letter(col) + "105")].value #Outlet6
			if(curr != None): 
				sumVal += curr
			curr = worksheet[(get_column_letter(col) + "69")].value #Inlet1
			if(curr != None): 
				sumVal -= curr
			curr = worksheet[(get_column_letter(col) + "73")].value #Inlet2
			if(curr != None): 
				sumVal -= curr
			curr = worksheet[(get_column_letter(col) + "77")].value #Inlet3
			if(curr != None): 
				sumVal -= curr
			curr = worksheet[(get_column_letter(col) + "81")].value #Inlet4
			if(curr != None): 
				sumVal -= curr
			curr = worksheet[(get_column_letter(col) + "107")].value #Heat Duty
			if(curr != None): 
				sumVal -= curr
			writeCell = worksheet[(get_column_letter(col) + "110")] 
			if(abs(sumVal) >= -1):
				writeCell.value = round_val(sumVal) #Write Cell - i.e Entropy generation
			else:	
				writeCell.value = round_val(sumVal) #Manual bypass, should throw error according to specs
				print("Sgen_Error, Entropy Balance is: " + str(sumVal))
				print("See block: " + current_block_name)
	
	mod_radfrac(worksheet, large_arr)
	add_temperature(worksheet)
	return 1

def mod_radfrac(worksheet, large_arr):
	rad_arr = []
	for arr in large_arr:
		current_block_name = (worksheet[(get_column_letter(arr[0]-1) + "2")].value)
		if(current_block_name == "RadFrac"):
			rad_arr = arr

	for col in worksheet.iter_cols(min_col=rad_arr[0]-1,max_col=rad_arr[0]-1, min_row=111, max_row=111):
		for cell in col:
			cell.value = "Heat In, MW"
			cell.fill = GREEN_HIGHLIGHT

	for col in worksheet.iter_cols(min_col=rad_arr[0]-1,max_col=rad_arr[0]-1, min_row=112, max_row=112):
		for cell in col:
			cell.value = "Heat Out, MW"
			cell.fill = GREEN_HIGHLIGHT

	for col in worksheet.iter_cols(min_col=rad_arr[0],max_col=rad_arr[len(rad_arr)-1], min_row=33, max_row=33): #Reboiler heat
		for cell in col:
			temp = cell.value
			cell.offset(row=111-33).value = temp

	for col in worksheet.iter_cols(min_col=rad_arr[0],max_col=rad_arr[len(rad_arr)-1], min_row=27, max_row=27): #Condensor heat duty
		for cell in col:
			temp = cell.value
			cell.offset(row=112-27).value = temp

	for col in worksheet.iter_cols(min_col=rad_arr[0],max_col=rad_arr[len(rad_arr)-1], min_row=110, max_row=110): 
		for cell in col:
			val1_off = cell.offset(row=32-110).value + 273.15
			val1 = cell.offset(row=1).value / val1_off
			val2_off = cell.offset(row=25-110).value + 273.15
			val2 = cell.offset(row=2).value / val2_off
			temp = cell.value - (val1 + val2) * 1000
			cell.value = temp


def add_temperature(worksheet):
	tempChangeFlag = 0
	temp_row = find_row_with_key(worksheet, "Calculated temperature [C]")
	if(temp_row == 0):
		tempChangeFlag = 1
		temp_row = find_row_with_key(worksheet, "Calculated temperature [K]")
	entropy_row = find_row_with_key(worksheet,"Entropy Generation kW/K")
	heat_row = find_row_with_key(worksheet,"Heat MW")

	worksheet['A111'].value = "Temperature, K"
	worksheet['A111'].fill = GREEN_HIGHLIGHT #Constant here fine

	for col in worksheet.iter_cols(min_col=2,max_col=5, min_row=temp_row, max_row=temp_row):
		for cell in col:
			if not (isinstance(cell.value, str)) and cell.value != None:
				if(tempChangeFlag):
					temp = cell.value
				else:
					temp = cell.value + 273.15
				cell.offset(row=111 - temp_row).value = temp

	for col in worksheet.iter_cols(min_col=2,max_col=5, min_row=entropy_row, max_row=entropy_row):
		for cell in col:
			temp = cell.value
			temperature =cell.offset(row=1).value * 1000
			cell.value = temp / temperature


def main():
	inputData = get_config_variables()
	streamWorkbook = inputData["streamBookName"]
	print("Working on: " + str(streamWorkbook))
	wb_stream = openpyxl.load_workbook(streamWorkbook)
	modifiedWS = copy_worksheet(wb_stream, "Aspen Data Tables Modified")
	with tqdm(total=100, file=sys.stdout) as pbar:
		for i in range(1):
			#Begin work on streams workbook
			step_six(modifiedWS)
			pbar.update(25)
			step_seven(modifiedWS, inputData["streamTitle"]) 
			wb_stream.save(streamWorkbook)
			overall = wb_stream.copy_worksheet(modifiedWS)
			overall.title = "Overall"
			pbar.update(25)
			step_eight(overall)
			wb_stream.save(streamWorkbook)
			pbar.update(25)
			check = step_nine(overall)
			wb_stream.save(streamWorkbook)
			if(check == 0):
				sys.exit()
			pbar.update(25)
	
	#Begin work on models workbook 
	inputData = get_config_variables()
	modelWorkbook = inputData["modelBookName"]
	print("Working on:" + str(modelWorkbook))
	wb_block = openpyxl.load_workbook(modelWorkbook)
	overallTitle = inputData["modelTitle"]
	overallWS = copy_worksheet(wb_block, "Overall")
	overall_text_add = inputData["overall_text_add"]
	with tqdm(total=100, file=sys.stdout) as pbar:
		for i in range(1):
			add_text(overallWS, overallTitle, overall_text_add)
			wb_block.save(modelWorkbook)
			pbar.update(25)
			inlet_array = prepare_for_overall_inlet(modifiedWS)
			inlet_vals_array = prepare_for_overall_inlet_vals(modifiedWS)
			step_twelve_inlet(overallWS, inlet_array, inlet_vals_array)
			pbar.update(25)
			outlet_array = prepare_for_overall_outlet(modifiedWS)
			outlet_vals_array = prepare_for_overall_outlet_vals(modifiedWS)
			step_twelve_outlet(overallWS, outlet_array, outlet_vals_array)
			pbar.update(25)
			step_thirteen(overallWS)
			check = step_fourteen(overallWS)
			wb_block.save(modelWorkbook)
			if(check == 0):
				sys.exit()
			pbar.update(25)
	print("Successful completion of program, please open the provided workbooks to see modifications\n")
if __name__ == '__main__':
	main()